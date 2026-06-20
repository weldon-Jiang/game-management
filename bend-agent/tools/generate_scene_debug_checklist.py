#!/usr/bin/env python3
"""
从 scene_schemas.py + profile_name_reader.py 生成场景/模板/OCR 调试 Excel 清单。

用法:
  cd bend-agent
  python tools/generate_scene_debug_checklist.py
  python tools/generate_scene_debug_checklist.py -o docs/scene_debug_checklist.xlsx
"""

from __future__ import annotations

import argparse
from collections import defaultdict
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
import sys

sys.path.insert(0, str(ROOT))

from configs.scene_schemas import SCENE_COLUMNS, SCENE_NAMES, SCENE_SCHEMAS
from src.agent.vision.template_manager import STEP4_REQUIRED_SCENE_IDS

# 自动化阶段 → 何时需要调坐标
PHASE_BY_SCENE: Dict[int, Tuple[str, str]] = {}
for sid, phase, when in [
    (1, "P0-串流就绪", "Step3 完成后首次识别 Xbox 主页；串流刚连上"),
    (2, "P0-串流就绪", "西瓜主页图标识别失败时"),
    (24, "P0-串流就绪", "小键盘/主页混合态识别"),
    (203, "P1-FC启动", "FC 磁贴识别；主页 OCR 跳过切换失败时"),
    (3, "P2-账号切换", "进入「档案和系统」失败时"),
    (4, "P2-账号切换", "档案子页识别失败时"),
    (5, "P2-账号切换", "「添加和切换」页识别失败时"),
    (6, "P2-账号切换", "选用户列表识别/OCR 定位档案失败时"),
    (7, "P2-账号切换", "「您希望做什么」弹窗识别失败时"),
    (8, "P2-账号切换", "「您希望做什么」变体识别失败时"),
    (9, "P2-账号切换", "「您希望做什么」变体识别失败时"),
    (10, "P3-登录键盘", "Microsoft 账号登录页识别失败时"),
    (101, "P4-进游戏", "FC 开场/加载识别失败时"),
    (126, "P4-进游戏", "「登陆游戏」识别失败时"),
    (127, "P4-进游戏", "UT 登录确认；MAIN_MENU 判定"),
    (147, "P5-UT主菜单", "UT 主页识别；MAIN_MENU 判定"),
    (149, "P5-UT主菜单", "「开始游戏」入口识别失败时"),
    (156, "P6-比赛导航", "Squad Battles 入口识别失败时"),
    (163, "P6-比赛导航", "SQB 子菜单识别失败时"),
    (166, "P6-比赛导航", "选对手页识别失败时"),
    (168, "P6-比赛导航", "对手格子（左上未打）识别失败时"),
    (176, "P6-比赛导航", "难度选择识别失败时"),
    (189, "P6-比赛导航", "开始比赛按钮识别失败时"),
    (191, "P6-比赛导航", "比赛结束识别失败时"),
    (204, "P9-异常", "长时间无操作断线弹窗识别失败时"),
]:
    PHASE_BY_SCENE[sid] = (phase, when)


def _phase_for_scene(scene_id: int) -> Tuple[str, str]:
    if scene_id in PHASE_BY_SCENE:
        return PHASE_BY_SCENE[scene_id]
    if 11 <= scene_id <= 64:
        return ("P3-登录键盘", f"虚拟键盘场景{scene_id}按键识别失败时")
    if 95 <= scene_id <= 99:
        return ("P7-新号添加", "新号添加游戏流程识别失败时")
    if 100 <= scene_id <= 146:
        return ("P8-EA首登", "EA/新号创建流程识别失败时")
    if 150 <= scene_id <= 201:
        return ("P6-比赛导航", "UT 内比赛/转会流程识别失败时")
    if 205 <= scene_id <= 255:
        return ("P8-EA首登", "EA 首登扩展场景识别失败时")
    if scene_id <= 9:
        return ("P0-串流就绪", "Xbox 系统 UI 识别失败时")
    return ("P9-其他", "对应自动化步骤报场景不匹配时")


OCR_REGIONS: List[Dict[str, Any]] = [
    {
        "region_id": "OCR-203-NAME",
        "scene_id": 203,
        "region_name": "主页左上角-显示名",
        "left": 32,
        "top": 14,
        "right": 280,
        "bottom": 52,
        "config_file": "src/agent/vision/profile_name_reader.py",
        "config_keys": "HOME203_NAME_LEFT/TOP/RIGHT/BOTTOM",
        "when": "主页 OCR 与 gamertag 不一致、debug_home_ocr_name_*.png 裁剪错位",
        "calibrate_tool": "scripts/debug/calibrate_home_ocr_regions.py --interactive",
        "phase": "P1-FC启动",
    },
    {
        "region_id": "OCR-203-EMAIL",
        "scene_id": 203,
        "region_name": "主页左上角-邮箱",
        "left": 32,
        "top": 44,
        "right": 400,
        "bottom": 82,
        "config_file": "src/agent/vision/profile_name_reader.py",
        "config_keys": "HOME203_EMAIL_LEFT/TOP/RIGHT/BOTTOM",
        "when": "邮箱行 OCR 为空或误识别、debug_home_ocr_email_*.png 错位",
        "calibrate_tool": "scripts/debug/calibrate_home_ocr_regions.py --interactive",
        "phase": "P1-FC启动",
    },
    {
        "region_id": "OCR-6-LIST",
        "scene_id": 6,
        "region_name": "档案列表整体",
        "left": 78,
        "top": 95,
        "right": 320,
        "bottom": 500,
        "config_file": "src/agent/vision/profile_name_reader.py",
        "config_keys": "SCENE6_LIST_LEFT/TOP/RIGHT/BOTTOM",
        "when": "场景6列表 OCR 兜底读不到任何 gamertag",
        "calibrate_tool": "scene_capture 截图 + 手工量坐标",
        "phase": "P2-账号切换",
    },
    {
        "region_id": "OCR-6-TEXT",
        "scene_id": 6,
        "region_name": "焦点行文字区",
        "left": 118,
        "top": "(动态: row_y±ROW_H/2)",
        "right": 300,
        "bottom": "(动态)",
        "config_file": "src/agent/vision/profile_name_reader.py",
        "config_keys": "SCENE6_TEXT_LEFT/RIGHT + SCENE6_ROW_HEIGHT",
        "when": "焦点行 gamertag OCR 错误但绿框检测正常",
        "calibrate_tool": "logs/debug_scene6_*.png 或 scene_capture",
        "phase": "P2-账号切换",
    },
    {
        "region_id": "OCR-6-GREEN",
        "scene_id": 6,
        "region_name": "绿框焦点检测带",
        "left": 78,
        "top": 95,
        "right": 86,
        "bottom": 500,
        "config_file": "src/agent/vision/profile_name_reader.py",
        "config_keys": "detect_focused_row_y() 扫描 SCENE6_LIST_* 左侧 8px",
        "when": "scene6_list_layout_present 为 false、无法定位焦点行",
        "calibrate_tool": "HSV 绿框阈值 _green_ratio()",
        "phase": "P2-账号切换",
    },
]

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)
PHASE_FILLS = {
    "P0-串流就绪": PatternFill("solid", fgColor="E2EFDA"),
    "P1-FC启动": PatternFill("solid", fgColor="DDEBF7"),
    "P2-账号切换": PatternFill("solid", fgColor="FFF2CC"),
    "P3-登录键盘": PatternFill("solid", fgColor="FCE4D6"),
    "P4-进游戏": PatternFill("solid", fgColor="E4DFEC"),
    "P5-UT主菜单": PatternFill("solid", fgColor="D9E1F2"),
    "P6-比赛导航": PatternFill("solid", fgColor="F8CBAD"),
    "P7-新号添加": PatternFill("solid", fgColor="EDEDED"),
    "P8-EA首登": PatternFill("solid", fgColor="EDEDED"),
    "P9-异常": PatternFill("solid", fgColor="F4CCCC"),
    "P9-其他": PatternFill("solid", fgColor="EDEDED"),
}
STEP4_FILL = PatternFill("solid", fgColor="C6EFCE")


def _style_header(ws, headers: List[str]) -> None:
    for col, title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def _autosize(ws, max_width: int = 48) -> None:
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        width = 10
        for row in range(1, min(ws.max_row, 200) + 1):
            val = ws.cell(row=row, column=col).value
            if val is None:
                continue
            width = max(width, min(max_width, len(str(val)) + 2))
        ws.column_dimensions[letter].width = width


def build_workbook() -> Workbook:
    wb = Workbook()

    # --- Sheet 1: 调试流程 ---
    ws_flow = wb.active
    ws_flow.title = "00-调试流程"
    flow_rows = [
        ["步骤", "做什么", "用哪张表", "产出/验证"],
        ["1", "从 scene_capture 或 logs 取 960×540 整帧截图", "-", "000_entry_initial.png 等"],
        ["2", "先查 Step4 必配场景是否缺模板", "03-Step4必配", "tools/verify_p0_readiness.py"],
        ["3", "模板匹配失败 → 调 template/search 坐标 + 重裁 PNG", "02-模板明细", "templates/{scene}.{tpl}.png"],
        ["4", "OCR 失败 → 调 OCR 区域常量", "04-OCR区域", "calibrate_home_ocr_regions.py"],
        ["5", "改 scene_schemas 后同步模板", "-", "python tools/sync_scene_schemas.py"],
        ["6", "改坐标后跑单场景测试", "-", "scripts/debug/test_home_ocr.py 等"],
        ["7", "在「调试状态」列打勾/备注", "02/04", "已验证/待验证"],
        [],
        ["坐标系", "960×540，原点左上角，X→右 Y→下", "", ""],
        ["模板区域", "裁切 PNG 用 template_left/top/right/bottom", "", ""],
        ["搜索区域", "帧内匹配范围 search_*，应包住模板区", "", ""],
        ["OCR 区域", "直接裁剪识别，无 likeness 阈值", "", ""],
    ]
    for r, row in enumerate(flow_rows, 1):
        for c, val in enumerate(row, 1):
            ws_flow.cell(row=r, column=c, value=val)
    _style_header(ws_flow, flow_rows[0])
    _autosize(ws_flow)

    # --- Sheet 2: 场景总览 ---
    ws_scene = wb.create_sheet("01-场景总览")
    scene_headers = [
        "序号",
        "自动化阶段",
        "场景ID",
        "场景名称",
        "模板数",
        "Step4必配",
        "识别类型",
        "何时需要调试",
        "改哪里(主)",
        "调试状态",
        "备注",
    ]
    _style_header(ws_scene, scene_headers)

    by_scene: Dict[int, List[dict]] = defaultdict(list)
    for schema in SCENE_SCHEMAS:
        row = dict(zip(SCENE_COLUMNS, schema))
        by_scene[int(row["scene_id"])].append(row)

    seq = 0
    for scene_id in sorted(by_scene.keys()):
        seq += 1
        phase, when = _phase_for_scene(scene_id)
        templates = by_scene[scene_id]
        name = SCENE_NAMES.get(scene_id, f"场景{scene_id}")
        is_step4 = "是" if scene_id in STEP4_REQUIRED_SCENE_IDS else ""
        ocr_note = ""
        if scene_id == 6:
            ocr_note = " + OCR选档"
        elif scene_id == 203:
            ocr_note = " + OCR档案名"
        values = [
            seq,
            phase,
            scene_id,
            name,
            len(templates),
            is_step4,
            f"模板匹配{ocr_note}",
            when,
            "configs/scene_schemas.py",
            "",
            "",
        ]
        r = seq + 1
        for c, val in enumerate(values, 1):
            cell = ws_scene.cell(row=r, column=c, value=val)
            if c == 2 and phase in PHASE_FILLS:
                cell.fill = PHASE_FILLS[phase]
            if c == 6 and is_step4:
                cell.fill = STEP4_FILL
    _autosize(ws_scene)

    # --- Sheet 3: 模板明细 ---
    ws_tpl = wb.create_sheet("02-模板明细")
    tpl_headers = [
        "序号",
        "自动化阶段",
        "场景ID",
        "场景名称",
        "模板ID",
        "模板文件",
        "template_left",
        "template_top",
        "template_right",
        "template_bottom",
        "search_id",
        "search_left",
        "search_top",
        "search_right",
        "search_bottom",
        "likeness",
        "algorithm",
        "Step4必配",
        "何时需要调试",
        "改哪里",
        "调试状态",
        "备注",
    ]
    _style_header(ws_tpl, tpl_headers)

    seq = 0
    for schema in SCENE_SCHEMAS:
        seq += 1
        row = dict(zip(SCENE_COLUMNS, schema))
        scene_id = int(row["scene_id"])
        template_id = int(row["template_id"])
        phase, when = _phase_for_scene(scene_id)
        name = SCENE_NAMES.get(scene_id, f"场景{scene_id}")
        is_step4 = "是" if scene_id in STEP4_REQUIRED_SCENE_IDS else ""
        tpl_file = f"{scene_id}.{template_id}.png"
        config_hint = (
            f"scene_schemas.py 搜 [{scene_id}, 960, 540, {template_id},"
        )
        values = [
            seq,
            phase,
            scene_id,
            name,
            template_id,
            tpl_file,
            row["template_left"],
            row["template_top"],
            row["template_right"],
            row["template_bottom"],
            row["search_id"],
            row["search_left"],
            row["search_top"],
            row["search_right"],
            row["search_bottom"],
            row["likeness"],
            row["algorithm"],
            is_step4,
            when,
            config_hint,
            "",
            "",
        ]
        r = seq + 1
        for c, val in enumerate(values, 1):
            cell = ws_tpl.cell(row=r, column=c, value=val)
            if c == 2 and phase in PHASE_FILLS:
                cell.fill = PHASE_FILLS[phase]
            if c == 18 and is_step4:
                cell.fill = STEP4_FILL
    _autosize(ws_tpl)

    # --- Sheet 4: Step4 必配 ---
    ws_p0 = wb.create_sheet("03-Step4必配")
    p0_headers = [
        "优先级",
        "场景ID",
        "场景名称",
        "模板ID",
        "模板文件",
        "template L/T/R/B",
        "search L/T/R/B",
        "likeness",
        "何时调试",
        "调试状态",
    ]
    _style_header(ws_p0, p0_headers)

    priority_map = {
        1: "1-主页", 2: "1-主页", 24: "1-主页",
        203: "2-FC磁贴",
        3: "3-切换", 4: "3-切换", 5: "3-切换", 6: "3-切换", 7: "3-切换",
        10: "4-登录",
        101: "5-进游戏", 126: "5-进游戏", 127: "5-进游戏",
        147: "6-UT", 149: "6-UT",
    }
    r = 1
    for schema in SCENE_SCHEMAS:
        row = dict(zip(SCENE_COLUMNS, schema))
        scene_id = int(row["scene_id"])
        if scene_id not in STEP4_REQUIRED_SCENE_IDS:
            continue
        r += 1
        template_id = int(row["template_id"])
        _, when = _phase_for_scene(scene_id)
        tpl_box = (
            f"{row['template_left']},{row['template_top']},"
            f"{row['template_right']},{row['template_bottom']}"
        )
        search_box = (
            f"{row['search_left']},{row['search_top']},"
            f"{row['search_right']},{row['search_bottom']}"
        )
        values = [
            priority_map.get(scene_id, "?"),
            scene_id,
            SCENE_NAMES.get(scene_id, ""),
            template_id,
            f"{scene_id}.{template_id}.png",
            tpl_box,
            search_box,
            row["likeness"],
            when,
            "",
        ]
        for c, val in enumerate(values, 1):
            ws_p0.cell(row=r, column=c, value=val)
    _autosize(ws_p0)

    # --- Sheet 5: OCR 区域 ---
    ws_ocr = wb.create_sheet("04-OCR区域")
    ocr_headers = [
        "序号",
        "自动化阶段",
        "区域ID",
        "场景ID",
        "场景名称",
        "区域说明",
        "left",
        "top",
        "right",
        "bottom",
        "改哪里(文件)",
        "改哪里(常量)",
        "何时需要调试",
        "校准工具",
        "调试状态",
        "备注",
    ]
    _style_header(ws_ocr, ocr_headers)

    for i, reg in enumerate(OCR_REGIONS, 1):
        scene_id = int(reg["scene_id"])
        values = [
            i,
            reg["phase"],
            reg["region_id"],
            scene_id,
            SCENE_NAMES.get(scene_id, ""),
            reg["region_name"],
            reg["left"],
            reg["top"],
            reg["right"],
            reg["bottom"],
            reg["config_file"],
            reg["config_keys"],
            reg["when"],
            reg["calibrate_tool"],
            "",
            "",
        ]
        r = i + 1
        for c, val in enumerate(values, 1):
            cell = ws_ocr.cell(row=r, column=c, value=val)
            if reg["phase"] in PHASE_FILLS:
                cell.fill = PHASE_FILLS[reg["phase"]]
    _autosize(ws_ocr)

    return wb


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate scene/template/OCR debug checklist Excel")
    default_out = ROOT / "docs" / "scene_debug_checklist.xlsx"
    parser.add_argument("-o", "--output", type=Path, default=default_out)
    args = parser.parse_args()

    args.output.parent.mkdir(parents=True, exist_ok=True)
    wb = build_workbook()
    wb.save(args.output)
    print(f"已生成: {args.output}")
    print(f"  场景数: {len(set(int(dict(zip(SCENE_COLUMNS, s))['scene_id']) for s in SCENE_SCHEMAS))}")
    print(f"  模板行: {len(SCENE_SCHEMAS)}")
    print(f"  OCR区域: {len(OCR_REGIONS)}")
    print(f"  Step4必配场景: {len(STEP4_REQUIRED_SCENE_IDS)}")


if __name__ == "__main__":
    main()
