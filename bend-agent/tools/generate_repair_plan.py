#!/usr/bin/env python3
"""
模板匹配修复计划：从 scene_schemas 生成/刷新清单，状态写入 template_repair_status.json。

用法:
  cd bend-agent

  # 生成/刷新 Markdown + Excel（保留已有状态）
  python tools/generate_repair_plan.py

  # 更新单个模板状态
  python tools/generate_repair_plan.py status --scene 6 --template 1 --set verified --note "conf=0.91"

  # 更新整场景自动化状态
  python tools/generate_repair_plan.py status --scene 6 --set ready --note "switch_to 通过"

  # 查看推荐下一项
  python tools/generate_repair_plan.py next
"""

from __future__ import annotations

import argparse
import json
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from configs.scene_schemas import SCENE_COLUMNS, SCENE_NAMES, SCENE_SCHEMAS
from src.agent.vision.template_manager import STEP4_REQUIRED_SCENE_IDS
from tools.automation_checkpoints import AUTOMATION_CHECKPOINTS

STATUS_JSON = ROOT / "docs" / "template_repair_status.json"
PLAN_MD = ROOT / "docs" / "TEMPLATE_REPAIR_PLAN.md"
PLAN_XLSX = ROOT / "docs" / "template_repair_plan.xlsx"

# 修复顺序：先 Step4 必配，再账号链，再进游戏/UT，再 SQB，最后其余
REPAIR_ORDER: List[int] = []
for group in [
    # 现代 Xbox 云串流首屏多为 203(FC主页)，场景1 为旧版顶栏布局，勿从 1 开始修
    [203, 1, 2, 24],
    [3, 4, 5, 6, 7, 8, 9, 10],
    list(range(11, 65)),
    [101, 126, 127],
    [147, 149],
    [156, 163, 164, 165, 166, 168, 169, 170, 171, 172, 173, 174, 175, 176, 177, 189, 191, 193, 194],
]:
    REPAIR_ORDER.extend(group)
_seen = set(REPAIR_ORDER)

PHASE_LABELS = {
    "P0": "P0 Step4必配（最高优先级）",
    "P1": "P1 账号切换链",
    "P2": "P2 登录虚拟键盘",
    "P3": "P3 FC 启动进游戏",
    "P4": "P4 UT 主菜单",
    "P5": "P5 SQB/比赛导航",
    "P6": "P6 其余场景",
}


def _phase_for_scene(scene_id: int) -> str:
    if scene_id in STEP4_REQUIRED_SCENE_IDS:
        return "P0"
    if scene_id in {2, 3, 4, 5, 6, 7, 8, 9, 10, 203}:
        return "P1"
    if 11 <= scene_id <= 64:
        return "P2"
    if scene_id in {101, 126, 127}:
        return "P3"
    if scene_id in {147, 149}:
        return "P4"
    if 150 <= scene_id <= 201:
        return "P5"
    return "P6"


def _scene_sort_key(scene_id: int) -> Tuple[int, int]:
    if scene_id in _seen:
        return (0, REPAIR_ORDER.index(scene_id))
    return (1, scene_id)


TEMPLATE_STATUS_LABEL = {
    "pending": "⬜ 待修复",
    "png_ok": "🟦 PNG已更新",
    "coords_ok": "🟨 坐标已调",
    "verified": "🟩 单模板验证通过",
    "skip": "⏸ 暂缓",
}

SCENE_STATUS_LABEL = {
    "pending": "⬜ 待开始",
    "partial": "🟡 部分通过",
    "ready": "🟢 场景拉通/自动化可运行",
    "skip": "⏸ 暂缓",
}


def _now() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def _default_template_status() -> Dict[str, Any]:
    return {"status": "pending", "note": "", "updated_at": ""}


def _default_scene_status() -> Dict[str, Any]:
    return {
        "status": "pending",
        "automation_ready": False,
        "note": "",
        "updated_at": "",
        "templates": {},
        "ocr": {},
    }


def _schema_rows_by_scene() -> Dict[int, List[Dict[str, Any]]]:
    grouped: Dict[int, List[Dict[str, Any]]] = defaultdict(list)
    for schema in SCENE_SCHEMAS:
        row = dict(zip(SCENE_COLUMNS, schema))
        grouped[int(row["scene_id"])].append(row)
    return grouped


def _ocr_regions_for_scene(scene_id: int) -> List[Dict[str, str]]:
    if scene_id == 203:
        return [
            {"id": "HOME203_NAME", "label": "主页显示名 OCR", "config": "HOME203_NAME_*"},
            {"id": "HOME203_EMAIL", "label": "主页邮箱 OCR", "config": "HOME203_EMAIL_*"},
        ]
    if scene_id == 6:
        return [
            {"id": "SCENE6_LIST", "label": "档案列表 OCR 区", "config": "SCENE6_LIST_*"},
            {"id": "SCENE6_TEXT", "label": "焦点行文字 OCR", "config": "SCENE6_TEXT_* + ROW_HEIGHT"},
        ]
    return []


def load_status() -> Dict[str, Any]:
    if STATUS_JSON.is_file():
        with open(STATUS_JSON, encoding="utf-8") as fh:
            data = json.load(fh)
    else:
        data = {"meta": {"version": 1, "last_updated": "", "notes": ""}, "scenes": {}}

    grouped = _schema_rows_by_scene()
    scenes: Dict[str, Any] = data.setdefault("scenes", {})

    for scene_id, rows in grouped.items():
        key = str(scene_id)
        scene = scenes.setdefault(key, _default_scene_status())
        tpl_map = scene.setdefault("templates", {})
        for row in rows:
            tid = str(int(row["template_id"]))
            tpl_map.setdefault(tid, _default_template_status())
        for ocr in _ocr_regions_for_scene(scene_id):
            scene.setdefault("ocr", {}).setdefault(
                ocr["id"], {"status": "pending", "note": "", "updated_at": ""}
            )

    data["meta"]["last_updated"] = _now()
    return data


def save_status(data: Dict[str, Any]) -> None:
    data["meta"]["last_updated"] = _now()
    STATUS_JSON.parent.mkdir(parents=True, exist_ok=True)
    with open(STATUS_JSON, "w", encoding="utf-8") as fh:
        json.dump(data, fh, ensure_ascii=False, indent=2)


def _count_progress(data: Dict[str, Any]) -> Dict[str, int]:
    grouped = _schema_rows_by_scene()
    total_tpl = sum(len(v) for v in grouped.values())
    verified_tpl = 0
    ready_scenes = 0
    partial_scenes = 0
    p0_scenes = set(str(s) for s in STEP4_REQUIRED_SCENE_IDS)

    for scene_id, rows in grouped.items():
        sk = str(scene_id)
        scene = data["scenes"].get(sk, {})
        st = scene.get("status", "pending")
        if st == "ready":
            ready_scenes += 1
        elif st == "partial":
            partial_scenes += 1
        tpl_map = scene.get("templates", {})
        for row in rows:
            tid = str(int(row["template_id"]))
            if tpl_map.get(tid, {}).get("status") == "verified":
                verified_tpl += 1

    p0_ready = sum(
        1 for sid in p0_scenes if data["scenes"].get(sid, {}).get("status") == "ready"
    )

    return {
        "total_scenes": len(grouped),
        "total_templates": total_tpl,
        "verified_templates": verified_tpl,
        "ready_scenes": ready_scenes,
        "partial_scenes": partial_scenes,
        "p0_total": len(p0_scenes),
        "p0_ready": p0_ready,
    }


def find_next_item(data: Dict[str, Any]) -> Optional[str]:
    grouped = _schema_rows_by_scene()
    order = REPAIR_ORDER + sorted(set(grouped) - set(REPAIR_ORDER))

    for scene_id in order:
        sk = str(scene_id)
        scene = data["scenes"].get(sk, {})
        if scene.get("status") == "ready":
            continue
        if scene.get("status") == "skip":
            continue

        rows = grouped.get(scene_id, [])
        tpl_map = scene.get("templates", {})
        for row in rows:
            tid = str(int(row["template_id"]))
            st = tpl_map.get(tid, {}).get("status", "pending")
            if st not in ("verified", "skip"):
                name = SCENE_NAMES.get(scene_id, f"scene{scene_id}")
                return f"场景 {scene_id} ({name}) → 模板 {tid} ({scene_id}.{tid}.png)"

        for ocr_id, ocr in scene.get("ocr", {}).items():
            if ocr.get("status") not in ("verified", "skip"):
                name = SCENE_NAMES.get(scene_id, f"scene{scene_id}")
                return f"场景 {scene_id} ({name}) → OCR {ocr_id}"

        if rows and all(
            tpl_map.get(str(int(r["template_id"])), {}).get("status") == "verified"
            for r in rows
        ):
            name = SCENE_NAMES.get(scene_id, f"scene{scene_id}")
            return f"场景 {scene_id} ({name}) → 全部模板已通过，待拉通验证（改场景状态为 ready）"

    return None


def cmd_status(args: argparse.Namespace) -> None:
    data = load_status()
    sk = str(args.scene)
    scene = data["scenes"].setdefault(sk, _default_scene_status())

    if args.template is not None:
        tid = str(args.template)
        tpl = scene.setdefault("templates", {}).setdefault(tid, _default_template_status())
        tpl["status"] = args.set
        tpl["note"] = args.note or tpl.get("note", "")
        tpl["updated_at"] = _now()
        print(f"已更新 场景{args.scene} 模板{args.template} → {args.set}")
    elif args.ocr:
        ocr = scene.setdefault("ocr", {}).setdefault(
            args.ocr, {"status": "pending", "note": "", "updated_at": ""}
        )
        ocr["status"] = args.set
        ocr["note"] = args.note or ocr.get("note", "")
        ocr["updated_at"] = _now()
        print(f"已更新 场景{args.scene} OCR {args.ocr} → {args.set}")
    else:
        scene["status"] = args.set
        scene["automation_ready"] = args.set == "ready"
        scene["note"] = args.note or scene.get("note", "")
        scene["updated_at"] = _now()
        print(f"已更新 场景{args.scene} → {args.set}")

    save_status(data)
    write_markdown(data)
    write_excel(data)
    nxt = find_next_item(data)
    if nxt:
        print(f"推荐下一项: {nxt}")


def write_markdown(data: Dict[str, Any]) -> None:
    grouped = _schema_rows_by_scene()
    prog = _count_progress(data)
    nxt = find_next_item(data)

    lines: List[str] = [
        "# 模板截图匹配修复计划",
        "",
        "> **状态文件**: `docs/template_repair_status.json`（改状态请用 CLI，勿手改 MD 表格）",
        "> **自动化卡点手册**: `docs/AUTOMATION_DEBUG_PLAYBOOK.md`",
        "> **跑完任务查卡哪**: `python tools/diagnose_last_run.py`",
        "> **刷新命令**: `python tools/generate_repair_plan.py`",
        "> **更新状态**: `python tools/generate_repair_plan.py status --scene 6 --template 1 --set verified`",
        "",
        f"最后刷新: {data['meta'].get('last_updated', '')}",
        "",
        "## 进度总览",
        "",
        "| 指标 | 数量 |",
        "|------|------|",
        f"| 场景总数 | {prog['total_scenes']} |",
        f"| 模板总数 | {prog['total_templates']} |",
        f"| 单模板已验证 | {prog['verified_templates']} / {prog['total_templates']} |",
        f"| 场景拉通 (ready) | {prog['ready_scenes']} |",
        f"| 场景部分通过 (partial) | {prog['partial_scenes']} |",
        f"| **P0 Step4必配 拉通** | **{prog['p0_ready']} / {prog['p0_total']}** |",
        "",
        "## 推荐下一项",
        "",
        f"**{nxt or '全部已完成 🎉'}**",
        "",
        "## 状态说明",
        "",
        "### 单模板状态",
        "",
        "| 状态码 | 显示 | 含义 |",
        "|--------|------|------|",
        "| `pending` | ⬜ 待修复 | 未开始 |",
        "| `png_ok` | 🟦 PNG已更新 | 小截图已重裁/替换 |",
        "| `coords_ok` | 🟨 坐标已调 | scene_schemas 坐标已改 |",
        "| `verified` | 🟩 单模板验证通过 | 对 scene_capture/debug 帧 matchTemplate 通过 |",
        "| `skip` | ⏸ 暂缓 | 本轮不修 |",
        "",
        "### 场景状态（拉通）",
        "",
        "| 状态码 | 显示 | 含义 |",
        "|--------|------|------|",
        "| `pending` | ⬜ 待开始 | 场景内仍有模板未 verified |",
        "| `partial` | 🟡 部分通过 | 部分模板 OK，自动化仍失败 |",
        "| `ready` | 🟢 场景拉通 | 该场景在自动化流程中稳定识别/可继续 |",
        "| `skip` | ⏸ 暂缓 | 本轮不修 |",
        "",
        "## 修复 SOP（每次碎片时间）",
        "",
        "1. **跑一遍自动化任务**（Step4 / 开始自动化）",
        "2. **查卡在哪**: `python tools/diagnose_last_run.py`",
        "3. 打开对应 `logs/debug_scene{N}_*.png` 整帧（缩放到 960×540）",
        "4. 用画图/PS 量 **template** 和 **search** 两个矩形，把 L,T,R,B 发我",
        "5. 我改 schema + 你重裁 `templates/{scene}.{tpl}.png` → 再跑任务验证",
        "6. 通过后: `python tools/generate_repair_plan.py status --scene N --template T --set verified`",
        "7. 整场景拉通: `--set ready`",
        "",
        "## 自动化卡点顺序（走任务时会按此经过）",
        "",
        "| 步骤 | 阶段 | 失败日志关键词 | 去哪找截图 | 你要发什么 |",
        "|------|------|----------------|------------|------------|",
    ]

    for cp in sorted(AUTOMATION_CHECKPOINTS, key=lambda x: x["order"]):
        scenes = ",".join(str(s) for s in cp.get("scenes", [])) or "-"
        lines.append(
            f"| **{cp['step']}** | {cp['phase']} | {cp.get('fail_log', '')} | "
            f"`{cp.get('debug_where', '')}` | {cp.get('you_do', '')} |"
        )

    lines.extend([
        "",
        "详细说明见 **docs/AUTOMATION_DEBUG_PLAYBOOK.md**",
        "",
        "---",
        "",
    ])

    by_phase: Dict[str, List[int]] = defaultdict(list)
    for scene_id in grouped:
        by_phase[_phase_for_scene(scene_id)].append(scene_id)

    for phase_key in ["P0", "P1", "P2", "P3", "P4", "P5", "P6"]:
        scene_ids = sorted(by_phase.get(phase_key, []), key=_scene_sort_key)
        if not scene_ids:
            continue
        lines.append(f"## {PHASE_LABELS[phase_key]}")
        lines.append("")

        for scene_id in scene_ids:
            sk = str(scene_id)
            scene = data["scenes"].get(sk, _default_scene_status())
            name = SCENE_NAMES.get(scene_id, f"场景{scene_id}")
            scene_st = SCENE_STATUS_LABEL.get(scene.get("status", "pending"), scene.get("status", ""))
            step4 = " **[Step4必配]**" if scene_id in STEP4_REQUIRED_SCENE_IDS else ""
            note = scene.get("note", "")
            updated = scene.get("updated_at", "")

            lines.append(f"### 场景 {scene_id} — {name}{step4}")
            lines.append("")
            lines.append(f"- **场景状态**: {scene_st}")
            if note:
                lines.append(f"- **场景备注**: {note}")
            if updated:
                lines.append(f"- **场景更新**: {updated}")
            lines.append("")
            lines.append("| 模板 | 文件 | template L/T/R/B | search L/T/R/B | likeness | 模板状态 | 备注 | 更新时间 |")
            lines.append("|------|------|------------------|----------------|----------|----------|------|----------|")

            for row in grouped[scene_id]:
                tid = str(int(row["template_id"]))
                tpl = scene.get("templates", {}).get(tid, _default_template_status())
                tpl_st = TEMPLATE_STATUS_LABEL.get(tpl.get("status", "pending"), tpl.get("status", ""))
                t_box = f"{row['template_left']},{row['template_top']},{row['template_right']},{row['template_bottom']}"
                s_box = f"{row['search_left']},{row['search_top']},{row['search_right']},{row['search_bottom']}"
                lines.append(
                    f"| {tid} | `{scene_id}.{tid}.png` | {t_box} | {s_box} | {row['likeness']} | {tpl_st} | {tpl.get('note', '')} | {tpl.get('updated_at', '')} |"
                )

            ocr_regions = _ocr_regions_for_scene(scene_id)
            if ocr_regions:
                lines.append("")
                lines.append("**OCR 区域（非模板 PNG）**")
                lines.append("")
                lines.append("| OCR ID | 说明 | 配置常量 | 状态 | 备注 | 更新时间 |")
                lines.append("|--------|------|----------|------|------|----------|")
                for ocr in ocr_regions:
                    ocr_st = scene.get("ocr", {}).get(ocr["id"], {})
                    st_label = TEMPLATE_STATUS_LABEL.get(
                        ocr_st.get("status", "pending"), ocr_st.get("status", "")
                    )
                    lines.append(
                        f"| {ocr['id']} | {ocr['label']} | `{ocr['config']}` | {st_label} | {ocr_st.get('note', '')} | {ocr_st.get('updated_at', '')} |"
                    )

            lines.append("")
            lines.append("---")
            lines.append("")

    PLAN_MD.parent.mkdir(parents=True, exist_ok=True)
    PLAN_MD.write_text("\n".join(lines), encoding="utf-8")
    print(f"Markdown: {PLAN_MD}")


def write_excel(data: Dict[str, Any]) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    grouped = _schema_rows_by_scene()
    prog = _count_progress(data)
    nxt = find_next_item(data)

    wb = Workbook()

    # Sheet 1: 进度
    ws0 = wb.active
    ws0.title = "00-进度与下一项"
    ws0.append(["指标", "值"])
    ws0.append(["最后刷新", data["meta"].get("last_updated", "")])
    ws0.append(["场景总数", prog["total_scenes"]])
    ws0.append(["模板总数", prog["total_templates"]])
    ws0.append(["单模板已验证", f"{prog['verified_templates']}/{prog['total_templates']}"])
    ws0.append(["场景拉通 ready", prog["ready_scenes"]])
    ws0.append(["P0 Step4必配拉通", f"{prog['p0_ready']}/{prog['p0_total']}"])
    ws0.append([])
    ws0.append(["推荐下一项", nxt or "全部完成"])
    ws0.append([])
    ws0.append(["更新命令示例", "python tools/generate_repair_plan.py status --scene 6 --template 1 --set verified"])

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)

    # Sheet 2: 修复清单（按优先级排序）
    ws = wb.create_sheet("01-修复清单")
    headers = [
        "修复优先级",
        "阶段",
        "Step4必配",
        "场景ID",
        "场景名称",
        "场景状态",
        "场景自动化ready",
        "类型",
        "模板/OCR ID",
        "文件/配置",
        "template/search 或 OCR",
        "likeness",
        "模板/OCR状态",
        "备注",
        "更新时间",
    ]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = header_fill
        cell.font = header_font
    ws.freeze_panes = "A2"

    order = REPAIR_ORDER + sorted(set(grouped) - set(REPAIR_ORDER))
    row_idx = 2
    priority = 0
    for scene_id in order:
        sk = str(scene_id)
        scene = data["scenes"].get(sk, _default_scene_status())
        name = SCENE_NAMES.get(scene_id, f"场景{scene_id}")
        phase = _phase_for_scene(scene_id)
        step4 = "是" if scene_id in STEP4_REQUIRED_SCENE_IDS else ""
        scene_st = scene.get("status", "pending")

        for row in grouped.get(scene_id, []):
            priority += 1
            tid = str(int(row["template_id"]))
            tpl = scene.get("templates", {}).get(tid, _default_template_status())
            t_box = (
                f"T:{row['template_left']},{row['template_top']},"
                f"{row['template_right']},{row['template_bottom']} "
                f"S:{row['search_left']},{row['search_top']},"
                f"{row['search_right']},{row['search_bottom']}"
            )
            ws.append([
                priority,
                phase,
                step4,
                scene_id,
                name,
                scene_st,
                "是" if scene.get("automation_ready") else "否",
                "模板",
                tid,
                f"{scene_id}.{tid}.png",
                t_box,
                row["likeness"],
                tpl.get("status", "pending"),
                tpl.get("note", ""),
                tpl.get("updated_at", ""),
            ])
            row_idx += 1

        for ocr in _ocr_regions_for_scene(scene_id):
            priority += 1
            ocr_st = scene.get("ocr", {}).get(ocr["id"], {})
            ws.append([
                priority,
                phase,
                step4,
                scene_id,
                name,
                scene_st,
                "是" if scene.get("automation_ready") else "否",
                "OCR",
                ocr["id"],
                ocr["config"],
                "profile_name_reader.py",
                "-",
                ocr_st.get("status", "pending"),
                ocr_st.get("note", ""),
                ocr_st.get("updated_at", ""),
            ])
            row_idx += 1

    # Sheet 3: 场景汇总
    ws2 = wb.create_sheet("02-场景汇总")
    h2 = ["阶段", "Step4必配", "场景ID", "场景名称", "模板数", "已验证模板", "场景状态", "自动化ready", "场景备注", "更新时间"]
    for c, h in enumerate(h2, 1):
        cell = ws2.cell(row=1, column=c, value=h)
        cell.fill = header_fill
        cell.font = header_font
    ws2.freeze_panes = "A2"

    for scene_id in order:
        if scene_id not in grouped:
            continue
        sk = str(scene_id)
        scene = data["scenes"].get(sk, _default_scene_status())
        rows = grouped[scene_id]
        verified = sum(
            1 for r in rows
            if scene.get("templates", {}).get(str(int(r["template_id"])), {}).get("status") == "verified"
        )
        ws2.append([
            _phase_for_scene(scene_id),
            "是" if scene_id in STEP4_REQUIRED_SCENE_IDS else "",
            scene_id,
            SCENE_NAMES.get(scene_id, ""),
            len(rows),
            f"{verified}/{len(rows)}",
            scene.get("status", "pending"),
            "是" if scene.get("automation_ready") else "否",
            scene.get("note", ""),
            scene.get("updated_at", ""),
        ])

    # Sheet 4: 自动化卡点
    ws3 = wb.create_sheet("03-自动化卡点")
    h3 = [
        "顺序", "步骤", "自动化阶段", "代码触发点", "失败日志关键词",
        "涉及场景", "相关模板", "相关OCR", "截图去哪找", "失败时你要做什么", "坐标提交格式", "修复状态",
    ]
    for c, h in enumerate(h3, 1):
        cell = ws3.cell(row=1, column=c, value=h)
        cell.fill = header_fill
        cell.font = header_font
    ws3.freeze_panes = "A2"
    for cp in sorted(AUTOMATION_CHECKPOINTS, key=lambda x: x["order"]):
        ws3.append([
            cp["order"],
            cp["step"],
            cp["phase"],
            cp.get("trigger", ""),
            cp.get("fail_log", ""),
            ",".join(str(s) for s in cp.get("scenes", [])),
            cp.get("templates", ""),
            cp.get("ocr", ""),
            cp.get("debug_where", ""),
            cp.get("you_do", ""),
            cp.get("coord_format", ""),
            "",  # 用户可手填：已修复/进行中
        ])
    ws3.column_dimensions["J"].width = 48
    ws3.column_dimensions["K"].width = 32

    for ws_item in (ws, ws2, ws3):
        for col in range(1, ws_item.max_column + 1):
            ws_item.column_dimensions[get_column_letter(col)].width = 16
    ws.column_dimensions["K"].width = 40
    ws.column_dimensions["N"].width = 24

    wb.save(PLAN_XLSX)
    print(f"Excel: {PLAN_XLSX}")


def cmd_next(_: argparse.Namespace) -> None:
    data = load_status()
    nxt = find_next_item(data)
    prog = _count_progress(data)
    print(f"P0 拉通: {prog['p0_ready']}/{prog['p0_total']}")
    print(f"模板 verified: {prog['verified_templates']}/{prog['total_templates']}")
    print(f"推荐下一项: {nxt or '全部已完成'}")


def cmd_generate(_: argparse.Namespace) -> None:
    data = load_status()
    save_status(data)
    write_markdown(data)
    write_excel(data)
    nxt = find_next_item(data)
    prog = _count_progress(data)
    print(f"进度: 模板 {prog['verified_templates']}/{prog['total_templates']} | P0 {prog['p0_ready']}/{prog['p0_total']}")
    print(f"推荐下一项: {nxt or '全部已完成'}")


def main() -> None:
    parser = argparse.ArgumentParser(description="模板匹配修复计划生成与状态更新")
    sub = parser.add_subparsers(dest="cmd")

    p_gen = sub.add_parser("generate", help="刷新 MD + Excel（默认）")
    p_gen.set_defaults(func=cmd_generate)

    p_next = sub.add_parser("next", help="查看推荐下一项")
    p_next.set_defaults(func=cmd_next)

    p_st = sub.add_parser("status", help="更新模板/场景/OCR 状态")
    p_st.add_argument("--scene", type=int, required=True)
    p_st.add_argument("--template", type=int, default=None)
    p_st.add_argument("--ocr", type=str, default=None)
    p_st.add_argument(
        "--set",
        required=True,
        choices=["pending", "png_ok", "coords_ok", "verified", "skip", "partial", "ready"],
    )
    p_st.add_argument("--note", type=str, default="")
    p_st.set_defaults(func=cmd_status)

    args = parser.parse_args()
    if args.cmd is None:
        cmd_generate(args)
    else:
        args.func(args)


if __name__ == "__main__":
    main()
